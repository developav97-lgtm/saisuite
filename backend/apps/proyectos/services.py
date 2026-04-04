"""
SaiSuite — Proyectos: Services
TODA la lógica de negocio va aquí. Las views solo orquestan.
"""
import logging
from decimal import Decimal
from django.db import transaction
try:
    from weasyprint import HTML as WeasyHTML, CSS as WeasyCSS
    _WEASYPRINT_AVAILABLE = True
except ImportError:  # pragma: no cover
    _WEASYPRINT_AVAILABLE = False
from django.db.models import Avg, ExpressionWrapper, F, Sum, Count, Q, QuerySet
from django.db.models import DecimalField as DbDecimalField
from django.utils import timezone
from rest_framework.exceptions import ValidationError, PermissionDenied

from apps.proyectos.models import (
    Project, Phase, Task, ProjectStakeholder, AccountingDocument, Milestone,
    ProjectStatus, PhaseStatus, Activity, ProjectActivity, ModuleSettings,
    ActivityType, MeasurementMode, ProjectType,
)

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# Excepciones del módulo
# ──────────────────────────────────────────────

class ProyectoException(ValidationError):
    """Base para errores del módulo de proyectos."""


class TransicionEstadoInvalidaException(ProyectoException):
    pass


class PresupuestoExcedidoException(ProyectoException):
    pass


class ProyectoNoEditableException(ProyectoException):
    pass


# ──────────────────────────────────────────────
# Avance automático
# ──────────────────────────────────────────────

def calcular_avance_fase_desde_tareas(fase_id) -> Decimal:
    """
    Recalcula porcentaje_avance de la fase basándose en el progreso de sus tareas.
    Fórmula: promedio de porcentaje_completado de todas las tareas activas
    (excluye canceladas). Tareas in_progress con progreso parcial contribuyen
    correctamente a diferencia del conteo de completadas.
    Devuelve el porcentaje calculado.
    """
    from apps.proyectos.models import Task
    tareas = Task.all_objects.filter(
        fase_id=fase_id,
    ).exclude(estado='cancelled')

    resultado = tareas.aggregate(promedio=Avg('porcentaje_completado'))
    promedio = resultado.get('promedio')

    if promedio is None:
        pct = Decimal('0')
    else:
        pct = Decimal(str(promedio)).quantize(Decimal('0.01'))

    Phase.objects.filter(id=fase_id).update(porcentaje_avance=pct)
    logger.info(
        'Avance fase recalculado desde tareas',
        extra={'fase_id': str(fase_id), 'porcentaje': str(pct)},
    )
    return pct


def calcular_avance_fase(fase_id) -> Decimal:
    """
    Recalcula y persiste porcentaje_avance de la fase basándose en sus actividades.
    Fórmula: Σ(ejecutado × costo) / Σ(planificado × costo) × 100.
    Devuelve el porcentaje calculado.
    """
    agg = ProjectActivity.all_objects.filter(fase_id=fase_id).aggregate(
        planificado=Sum(ExpressionWrapper(
            F('cantidad_planificada') * F('costo_unitario'),
            output_field=DbDecimalField(),
        )),
        ejecutado=Sum(ExpressionWrapper(
            F('cantidad_ejecutada') * F('costo_unitario'),
            output_field=DbDecimalField(),
        )),
    )
    planificado = agg['planificado'] or Decimal('0')
    ejecutado   = agg['ejecutado']   or Decimal('0')

    if planificado > Decimal('0'):
        pct = min(
            (ejecutado / planificado * Decimal('100')).quantize(Decimal('0.01')),
            Decimal('100'),
        )
    else:
        pct = Decimal('0')

    Phase.objects.filter(id=fase_id).update(porcentaje_avance=pct)
    logger.info('Avance fase recalculado', extra={'fase_id': str(fase_id), 'porcentaje': str(pct)})
    return pct


def calcular_avance_proyecto(proyecto_id) -> Decimal:
    """
    Recalcula y persiste porcentaje_avance del proyecto como promedio de sus fases activas.
    Devuelve el porcentaje calculado.
    """
    result = Phase.objects.filter(proyecto_id=proyecto_id, activo=True).aggregate(
        avg=Avg('porcentaje_avance')
    )
    avg = result['avg']
    pct = (avg.quantize(Decimal('0.01')) if avg is not None else Decimal('0'))

    Project.objects.filter(id=proyecto_id).update(porcentaje_avance=pct)
    logger.info('Avance proyecto recalculado', extra={'proyecto_id': str(proyecto_id), 'porcentaje': str(pct)})
    return pct


# ──────────────────────────────────────────────
# Recálculo ActividadProyecto desde tareas
# ──────────────────────────────────────────────

def recalcular_cantidad_ejecutada_ap(actividad_proyecto_id) -> Decimal:
    """
    Recalcula cantidad_ejecutada y porcentaje_avance de una ActividadProyecto
    sumando el progreso de todas sus tareas asociadas.

    - Modo 'hora'/'horas': suma horas_registradas de las tareas
    - Otros: suma cantidad_registrada de las tareas
    """
    try:
        ap = ProjectActivity.objects.select_related('actividad').get(id=actividad_proyecto_id)
    except ProjectActivity.DoesNotExist:
        return Decimal('0')

    from apps.proyectos.models import Task
    tareas = Task.all_objects.filter(actividad_proyecto_id=actividad_proyecto_id)

    unidad = (ap.actividad.unidad_medida or '').lower().strip()
    if unidad in ('hora', 'horas'):
        total = tareas.aggregate(s=Sum('horas_registradas'))['s'] or Decimal('0')
    else:
        total = tareas.aggregate(s=Sum('cantidad_registrada'))['s'] or Decimal('0')

    cantidad = Decimal(str(total)).quantize(Decimal('0.01'))

    if ap.cantidad_planificada > Decimal('0'):
        pct = min((cantidad / ap.cantidad_planificada * Decimal('100')).quantize(Decimal('0.01')), Decimal('100'))
    else:
        pct = Decimal('0')

    ProjectActivity.objects.filter(id=actividad_proyecto_id).update(
        cantidad_ejecutada=cantidad,
        porcentaje_avance=pct,
    )
    logger.info(
        'ActividadProyecto recalculada desde tareas',
        extra={'ap_id': str(actividad_proyecto_id), 'cantidad': str(cantidad), 'pct': str(pct)},
    )
    return cantidad


# ──────────────────────────────────────────────
# ConfiguracionModuloService
# ──────────────────────────────────────────────

class ConfiguracionModuloService:

    @staticmethod
    def get_or_create(company) -> ModuleSettings:
        obj, _ = ModuleSettings.objects.get_or_create(company=company)
        return obj

    @staticmethod
    def update(company, data: dict) -> ModuleSettings:
        obj, _ = ModuleSettings.objects.get_or_create(company=company)
        for key, val in data.items():
            setattr(obj, key, val)
        obj.save()
        logger.info('ModuleSettings actualizada', extra={'company_id': str(company.id)})
        return obj


# ──────────────────────────────────────────────
# Máquina de estados
# ──────────────────────────────────────────────

# Transiciones válidas: estado_actual → [estados_destino]
TRANSICIONES_VALIDAS: dict[str, list[str]] = {
    ProjectStatus.DRAFT:     [ProjectStatus.PLANNED, ProjectStatus.CANCELLED],
    ProjectStatus.PLANNED:  [ProjectStatus.IN_PROGRESS, ProjectStatus.DRAFT, ProjectStatus.CANCELLED],
    ProjectStatus.IN_PROGRESS: [ProjectStatus.SUSPENDED, ProjectStatus.CLOSED, ProjectStatus.CANCELLED],
    ProjectStatus.SUSPENDED:   [ProjectStatus.IN_PROGRESS, ProjectStatus.CANCELLED],
    ProjectStatus.CLOSED:      [],
    ProjectStatus.CANCELLED:    [],
}

# Estados que permiten edición del proyecto
ESTADOS_EDITABLES = {ProjectStatus.DRAFT, ProjectStatus.PLANNED}


# ──────────────────────────────────────────────
# ProyectoService
# ──────────────────────────────────────────────

class ProyectoService:

    @staticmethod
    def list_proyectos(company=None) -> QuerySet:
        """
        Retorna QuerySet de proyectos activos con anotaciones útiles.
        `company` se pasa explícitamente desde la view para garantizar
        aislamiento multi-tenant con JWT (el middleware no intercepta DRF auth).
        """
        qs = (
            Project.all_objects
            .filter(activo=True)
            .select_related('gerente', 'coordinador', 'company')
            .annotate(fases_count=Count('phases', filter=Q(phases__activo=True)))
        )
        if company is not None:
            qs = qs.filter(company=company)
        return qs

    @staticmethod
    def get_proyecto(proyecto_id: str) -> Project:
        """Retorna proyecto con relaciones cargadas."""
        return (
            Project.objects
            .select_related('gerente', 'coordinador', 'company')
            .prefetch_related('phases')
            .get(id=proyecto_id)
        )

    @staticmethod
    def _generar_codigo(company) -> str:
        """Genera código autoincremental por empresa: PRY-001, PRY-002, ..."""
        codigos = (
            Project.all_objects
            .filter(company=company, codigo__startswith='PRY-')
            .values_list('codigo', flat=True)
        )
        max_num = 0
        for c in codigos:
            try:
                num = int(c.split('-')[1])
                if num > max_num:
                    max_num = num
            except (IndexError, ValueError):
                pass
        return f'PRY-{max_num + 1:03d}'

    @staticmethod
    def _validar_gerente(gerente_id: str, company) -> object:
        """Verifica que el gerente pertenece a la misma empresa."""
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            gerente = User.objects.get(id=gerente_id, company=company)
        except User.DoesNotExist:
            raise ProyectoException(
                {'gerente': 'El gerente debe pertenecer a la misma empresa.'}
            )
        return gerente

    @staticmethod
    def _validar_coordinador(coordinador_id: str | None, company) -> object | None:
        """Verifica que el coordinador pertenece a la misma empresa."""
        if not coordinador_id:
            return None
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            return User.objects.get(id=coordinador_id, company=company)
        except User.DoesNotExist:
            raise ProyectoException(
                {'coordinador': 'El coordinador debe pertenecer a la misma empresa.'}
            )

    @staticmethod
    def create_proyecto(data: dict, user) -> Project:
        """Crea un nuevo proyecto validando reglas de negocio."""
        company = user.company
        gerente_id     = data.pop('gerente')
        coordinador_id = data.pop('coordinador', None)

        gerente     = ProyectoService._validar_gerente(str(gerente_id), company)
        coordinador = ProyectoService._validar_coordinador(
            str(coordinador_id) if coordinador_id else None, company
        )

        # Auto-generar código: usar consecutivo seleccionado por el usuario, fallback a secuencial
        consecutivo_id = data.pop('consecutivo_id', None)
        if not data.get('codigo'):
            from apps.core.services import generar_consecutivo
            codigo = generar_consecutivo(str(consecutivo_id)) if consecutivo_id else None
            if not codigo:
                codigo = ProyectoService._generar_codigo(company)
            data['codigo'] = codigo

        proyecto = Project(
            company=company,
            gerente=gerente,
            coordinador=coordinador,
            **data,
        )
        proyecto.full_clean()
        proyecto.save()

        logger.info(
            'Proyecto creado',
            extra={'proyecto_id': str(proyecto.id), 'codigo': proyecto.codigo, 'user': user.email},
        )
        return proyecto

    @staticmethod
    def update_proyecto(proyecto: Project, data: dict) -> Project:
        """Actualiza proyecto — solo permitido en estados editables."""
        if proyecto.estado not in ESTADOS_EDITABLES:
            raise ProyectoNoEditableException(
                f'No se puede editar un proyecto en estado "{proyecto.get_estado_display()}".'
            )

        company = proyecto.company
        if 'gerente' in data:
            data['gerente'] = ProyectoService._validar_gerente(str(data['gerente']), company)
        if 'coordinador' in data:
            data['coordinador'] = ProyectoService._validar_coordinador(
                str(data['coordinador']) if data['coordinador'] else None, company
            )

        for campo, valor in data.items():
            setattr(proyecto, campo, valor)

        proyecto.full_clean()
        proyecto.save()

        logger.info('Proyecto actualizado', extra={'proyecto_id': str(proyecto.id)})
        return proyecto

    @staticmethod
    def soft_delete_proyecto(proyecto: Project) -> None:
        """Soft delete: marca proyecto y todas sus fases como inactivos."""
        proyecto.activo = False
        proyecto.save(update_fields=['activo', 'updated_at'])
        proyecto.phases.filter(activo=True).update(activo=False)
        logger.info('Proyecto eliminado (soft)', extra={'proyecto_id': str(proyecto.id)})

    @staticmethod
    def cambiar_estado(proyecto: Project, nuevo_estado: str, forzar: bool = False) -> Project:
        """
        Cambia el estado del proyecto con validación de transiciones y precondiciones.
        """
        estado_actual = proyecto.estado
        transiciones  = TRANSICIONES_VALIDAS.get(estado_actual, [])

        if nuevo_estado not in transiciones:
            raise TransicionEstadoInvalidaException(
                f'No se puede pasar de "{proyecto.get_estado_display()}" '
                f'a "{dict(ProjectStatus.choices).get(nuevo_estado, nuevo_estado)}".'
                if hasattr(proyecto, 'EstadoProyecto')
                else f'Transición de estado inválida: {estado_actual} → {nuevo_estado}.'
            )

        # Precondiciones por transición
        if estado_actual == ProjectStatus.DRAFT and nuevo_estado == ProjectStatus.PLANNED:
            fases_activas = proyecto.phases.filter(activo=True).count()
            if fases_activas == 0:
                raise TransicionEstadoInvalidaException(
                    'Para planificar el proyecto debe tener al menos 1 fase definida.'
                )
            if proyecto.presupuesto_total <= Decimal('0'):
                raise TransicionEstadoInvalidaException(
                    'Para planificar el proyecto debe tener un presupuesto total mayor a cero.'
                )

        if estado_actual == ProjectStatus.PLANNED and nuevo_estado == ProjectStatus.IN_PROGRESS:
            try:
                config = ModuleSettings.objects.get(company=proyecto.company)
                requiere_sync = config.requiere_sync_saiopen_para_ejecucion
            except ModuleSettings.DoesNotExist:
                requiere_sync = False
            if requiere_sync and not proyecto.sincronizado_con_saiopen:
                raise TransicionEstadoInvalidaException(
                    'El proyecto debe estar sincronizado con Saiopen antes de iniciar ejecución.'
                )

        if estado_actual == ProjectStatus.IN_PROGRESS and nuevo_estado == ProjectStatus.CLOSED:
            if not forzar:
                fases_incompletas = proyecto.phases.filter(
                    activo=True, porcentaje_avance__lt=100
                ).count()
                if fases_incompletas > 0:
                    raise TransicionEstadoInvalidaException(
                        f'Hay {fases_incompletas} fase(s) con avance menor al 100%. '
                        f'Use forzar=True para cerrar igualmente.'
                    )

        # Registrar fecha real de inicio/fin
        if nuevo_estado == ProjectStatus.IN_PROGRESS and not proyecto.fecha_inicio_real:
            proyecto.fecha_inicio_real = timezone.now().date()
        if nuevo_estado in {ProjectStatus.CLOSED, ProjectStatus.CANCELLED}:
            if not proyecto.fecha_fin_real:
                proyecto.fecha_fin_real = timezone.now().date()

        proyecto.estado = nuevo_estado
        proyecto.save(update_fields=['estado', 'fecha_inicio_real', 'fecha_fin_real', 'updated_at'])

        logger.info(
            'Estado de proyecto cambiado',
            extra={
                'proyecto_id': str(proyecto.id),
                'estado_anterior': estado_actual,
                'estado_nuevo': nuevo_estado,
            },
        )
        return proyecto

    @staticmethod
    def get_estado_financiero(proyecto: Project) -> dict:
        """
        Calcula el estado financiero del proyecto:
        presupuesto, ejecutado (docs contables), AIU y desviaciones.
        """
        from django.db.models import Sum as DbSum

        # Presupuesto por categorías (suma de fases activas)
        fases_agg = proyecto.phases.filter(activo=True).aggregate(
            total_mano_obra=DbSum('presupuesto_mano_obra'),
            total_materiales=DbSum('presupuesto_materiales'),
            total_subcontratos=DbSum('presupuesto_subcontratos'),
            total_equipos=DbSum('presupuesto_equipos'),
            total_otros=DbSum('presupuesto_otros'),
        )

        def _d(val) -> Decimal:
            return val or Decimal('0')

        presupuesto_costos = (
            _d(fases_agg['total_mano_obra'])
            + _d(fases_agg['total_materiales'])
            + _d(fases_agg['total_subcontratos'])
            + _d(fases_agg['total_equipos'])
            + _d(fases_agg['total_otros'])
        )

        # AIU sobre costos directos
        aiu_factor = (
            proyecto.porcentaje_administracion
            + proyecto.porcentaje_imprevistos
            + proyecto.porcentaje_utilidad
        ) / Decimal('100')
        precio_venta_aiu = presupuesto_costos * (1 + aiu_factor)

        # Costo ejecutado (documentos contables importados de Saiopen)
        docs_agg = proyecto.documents.aggregate(ejecutado=DbSum('valor_neto'))
        costo_ejecutado = _d(docs_agg['ejecutado'])

        # Avance financiero
        pct_financiero = (
            (costo_ejecutado / presupuesto_costos * 100)
            if presupuesto_costos > 0
            else Decimal('0')
        )

        # Avance físico (promedio ponderado de fases — simplificado por orden)
        fases_avance = proyecto.phases.filter(activo=True).values_list('porcentaje_avance', flat=True)
        pct_fisico = (
            sum(fases_avance) / len(fases_avance)
            if fases_avance
            else Decimal('0')
        )

        return {
            'presupuesto_total':       str(proyecto.presupuesto_total),
            'presupuesto_costos':      str(presupuesto_costos),
            'precio_venta_aiu':        str(precio_venta_aiu.quantize(Decimal('0.01'))),
            'costo_ejecutado':         str(costo_ejecutado),
            'porcentaje_avance_fisico':  str(round(pct_fisico, 2)),
            'porcentaje_avance_financiero': str(round(pct_financiero, 2)),
            'desviacion_presupuesto':  str((costo_ejecutado - presupuesto_costos).quantize(Decimal('0.01'))),
            'desglose_presupuesto': {
                'mano_obra':    str(_d(fases_agg['total_mano_obra'])),
                'materiales':   str(_d(fases_agg['total_materiales'])),
                'subcontratos': str(_d(fases_agg['total_subcontratos'])),
                'equipos':      str(_d(fases_agg['total_equipos'])),
                'otros':        str(_d(fases_agg['total_otros'])),
            },
            'aiu': {
                'porcentaje_administracion': str(proyecto.porcentaje_administracion),
                'porcentaje_imprevistos':    str(proyecto.porcentaje_imprevistos),
                'porcentaje_utilidad':       str(proyecto.porcentaje_utilidad),
                'valor_aiu':                 str((precio_venta_aiu - presupuesto_costos).quantize(Decimal('0.01'))),
            },
        }


# ──────────────────────────────────────────────
# FaseService
# ──────────────────────────────────────────────

class FaseService:

    @staticmethod
    def list_fases(proyecto: Project) -> QuerySet:
        """Retorna fases activas del proyecto, ordenadas."""
        return proyecto.phases.filter(activo=True).order_by('orden')

    @staticmethod
    def _calcular_presupuesto_fases(proyecto: Project, excluir_fase_id=None) -> Decimal:
        """Suma el presupuesto total de todas las fases activas del proyecto."""
        qs = proyecto.phases.filter(activo=True)
        if excluir_fase_id:
            qs = qs.exclude(id=excluir_fase_id)

        agg = qs.aggregate(
            total=Sum('presupuesto_mano_obra')
                + Sum('presupuesto_materiales')
                + Sum('presupuesto_subcontratos')
                + Sum('presupuesto_equipos')
                + Sum('presupuesto_otros')
        )
        return agg.get('total') or Decimal('0')

    @staticmethod
    @transaction.atomic
    def create_fase(proyecto: Project, data: dict) -> Phase:
        """Crea una fase validando presupuesto y orden."""
        if proyecto.estado in {ProjectStatus.CLOSED, ProjectStatus.CANCELLED}:
            raise ProyectoNoEditableException(
                f'No se pueden agregar fases a un proyecto en estado "{proyecto.get_estado_display()}".'
            )

        # Bloquear fila del proyecto para evitar race conditions
        proyecto_locked = Project.all_objects.select_for_update().get(id=proyecto.id)

        # Calcular presupuesto de nueva fase
        nueva_fase_presupuesto = (
            (data.get('presupuesto_mano_obra') or Decimal('0'))
            + (data.get('presupuesto_materiales') or Decimal('0'))
            + (data.get('presupuesto_subcontratos') or Decimal('0'))
            + (data.get('presupuesto_equipos') or Decimal('0'))
            + (data.get('presupuesto_otros') or Decimal('0'))
        )

        presupuesto_actual = FaseService._calcular_presupuesto_fases(proyecto_locked)
        if (presupuesto_actual + nueva_fase_presupuesto) > proyecto_locked.presupuesto_total:
            disponible = proyecto_locked.presupuesto_total - presupuesto_actual
            raise PresupuestoExcedidoException(
                f'El presupuesto de las fases excedería el presupuesto total del proyecto. '
                f'Disponible: {disponible}. Solicitado: {nueva_fase_presupuesto}.'
            )

        # Auto-ordenar si no viene orden
        if 'orden' not in data or data['orden'] is None:
            ultimo_orden = (
                proyecto_locked.phases.filter(activo=True)
                .order_by('-orden')
                .values_list('orden', flat=True)
                .first()
            )
            data['orden'] = (ultimo_orden or 0) + 1

        fase = Phase(
            proyecto=proyecto_locked,
            company=proyecto_locked.company,
            **data,
        )
        fase.full_clean()
        fase.save()

        logger.info(
            'Fase creada',
            extra={'fase_id': str(fase.id), 'proyecto_id': str(proyecto.id), 'orden': fase.orden},
        )
        return fase

    @staticmethod
    @transaction.atomic
    def update_fase(fase: Phase, data: dict) -> Phase:
        """Actualiza una fase validando presupuesto."""
        proyecto_locked = Project.all_objects.select_for_update().get(id=fase.proyecto_id)

        # Calcular nuevo presupuesto de la fase
        nueva_mano_obra    = data.get('presupuesto_mano_obra',    fase.presupuesto_mano_obra)
        nueva_materiales   = data.get('presupuesto_materiales',   fase.presupuesto_materiales)
        nueva_subcontratos = data.get('presupuesto_subcontratos', fase.presupuesto_subcontratos)
        nueva_equipos      = data.get('presupuesto_equipos',      fase.presupuesto_equipos)
        nueva_otros        = data.get('presupuesto_otros',        fase.presupuesto_otros)

        nuevo_presupuesto_fase = (
            nueva_mano_obra + nueva_materiales
            + nueva_subcontratos + nueva_equipos + nueva_otros
        )

        presupuesto_resto = FaseService._calcular_presupuesto_fases(
            proyecto_locked, excluir_fase_id=fase.id
        )
        if (presupuesto_resto + nuevo_presupuesto_fase) > proyecto_locked.presupuesto_total:
            disponible = proyecto_locked.presupuesto_total - presupuesto_resto
            raise PresupuestoExcedidoException(
                f'El presupuesto actualizado excedería el presupuesto total del proyecto. '
                f'Disponible: {disponible}. Solicitado: {nuevo_presupuesto_fase}.'
            )

        for campo, valor in data.items():
            setattr(fase, campo, valor)

        fase.full_clean()
        fase.save()

        logger.info('Fase actualizada', extra={'fase_id': str(fase.id)})
        return fase

    @staticmethod
    @transaction.atomic
    def activar_fase(fase: Phase) -> Phase:
        """
        Activa una fase (estado → activa).
        Solo puede haber una fase activa por proyecto a la vez.
        La fase anterior activa pasa a 'planificada' automáticamente.
        """
        if fase.estado == PhaseStatus.COMPLETED:
            raise ProyectoException('No se puede activar una fase ya completada.')
        if fase.estado == PhaseStatus.CANCELLED:
            raise ProyectoException('No se puede activar una fase cancelada.')

        # Desactivar cualquier otra fase activa del mismo proyecto
        Phase.objects.filter(
            proyecto=fase.proyecto,
            estado=PhaseStatus.ACTIVE,
        ).exclude(id=fase.id).update(
            estado=PhaseStatus.PLANNED,
        )

        fase.estado = PhaseStatus.ACTIVE
        if not fase.fecha_inicio_real:
            from django.utils import timezone
            fase.fecha_inicio_real = timezone.now().date()
        fase.save(update_fields=['estado', 'fecha_inicio_real', 'updated_at'])

        logger.info('Fase activada', extra={'fase_id': str(fase.id), 'proyecto_id': str(fase.proyecto_id)})
        return fase

    @staticmethod
    @transaction.atomic
    def completar_fase(fase: Phase) -> Phase:
        """Marca una fase como completada y registra fecha real de fin."""
        if fase.estado not in (PhaseStatus.ACTIVE, PhaseStatus.PLANNED):
            raise ProyectoException(
                f'Solo se pueden completar fases activas o planificadas. Estado actual: {fase.estado}.'
            )

        from django.utils import timezone
        fase.estado = PhaseStatus.COMPLETED
        if not fase.fecha_fin_real:
            fase.fecha_fin_real = timezone.now().date()
        fase.save(update_fields=['estado', 'fecha_fin_real', 'updated_at'])

        # Recalcular avance del proyecto
        calcular_avance_proyecto(fase.proyecto_id)

        logger.info('Fase completada', extra={'fase_id': str(fase.id)})
        return fase

    @staticmethod
    def soft_delete_fase(fase: Phase) -> None:
        """Soft delete de la fase."""
        fase.activo = False
        fase.save(update_fields=['activo', 'updated_at'])
        logger.info('Fase eliminada (soft)', extra={'fase_id': str(fase.id)})

    @staticmethod
    @transaction.atomic
    def reorder_fases(proyecto: Project, ordered_ids: list) -> QuerySet:
        """
        Reasigna el orden de las fases activas de un proyecto.

        ordered_ids: lista de UUIDs en el nuevo orden deseado (posición 0 → orden 1).

        El unique_together(proyecto, orden) está marcado DEFERRABLE INITIALLY DEFERRED
        (migración 0021), por lo que PostgreSQL valida la unicidad al final de la
        transacción y no fila a fila. Esto permite hacer los UPDATEs secuenciales
        dentro del mismo atomic() sin conflictos intermedios.
        """
        fases_qs = Phase.all_objects.select_for_update().filter(
            proyecto=proyecto, activo=True
        )
        fases_por_id = {str(f.id): f for f in fases_qs}

        for fid in ordered_ids:
            if str(fid) not in fases_por_id:
                raise ProyectoException(
                    f'La fase {fid} no pertenece a este proyecto o está inactiva.'
                )

        for i, fid in enumerate(ordered_ids):
            Phase.all_objects.filter(id=fid).update(orden=i + 1)

        logger.info(
            'Fases reordenadas',
            extra={'proyecto_id': str(proyecto.id), 'nuevo_orden': [str(x) for x in ordered_ids]},
        )
        return FaseService.list_fases(proyecto)


# ──────────────────────────────────────────────
# TerceroProyectoService
# ──────────────────────────────────────────────

class TerceroProyectoService:

    @staticmethod
    def list_terceros(proyecto: Project, fase_id: str | None = None) -> QuerySet:
        """Retorna terceros activos del proyecto. Filtra por fase si se provee fase_id."""
        qs = (
            ProjectStakeholder.all_objects
            .filter(proyecto=proyecto, activo=True)
            .select_related('fase', 'tercero_fk')
            .order_by('rol', 'tercero_nombre')
        )
        if fase_id:
            qs = qs.filter(fase_id=fase_id)
        return qs

    @staticmethod
    @transaction.atomic
    def vincular_tercero(proyecto: Project, data: dict) -> ProjectStakeholder:
        """
        Vincula un tercero al proyecto.
        Regla: un tercero puede tener múltiples roles en el mismo proyecto,
        pero no el mismo rol+fase duplicado (unique_together).
        """
        # Validación explícita de duplicado (unique_together no captura NULL en SQL)
        if ProjectStakeholder.all_objects.filter(
            proyecto=proyecto,
            tercero_id=data.get('tercero_id'),
            rol=data.get('rol'),
            fase=data.get('fase'),
        ).exists():
            raise ValidationError(
                {'non_field_errors': 'Este tercero ya tiene el mismo rol en esta fase.'}
            )

        tercero = ProjectStakeholder(
            proyecto=proyecto,
            company=proyecto.company,
            **data,
        )
        try:
            tercero.full_clean()
            tercero.save()
        except Exception as exc:
            raise ValidationError(
                {'non_field_errors': f'Este tercero ya tiene el mismo rol en esta fase: {exc}'}
            )

        logger.info(
            'Tercero vinculado al proyecto',
            extra={
                'proyecto_id': str(proyecto.id),
                'tercero_id': data.get('tercero_id'),
                'rol': data.get('rol'),
            },
        )
        return tercero

    @staticmethod
    def desvincular_tercero(tercero: ProjectStakeholder) -> None:
        """Soft delete del tercero vinculado."""
        tercero.activo = False
        tercero.save(update_fields=['activo', 'updated_at'])
        logger.info(
            'Tercero desvinculado del proyecto',
            extra={'tercero_proyecto_id': str(tercero.id)},
        )


# ──────────────────────────────────────────────
# DocumentoContableService
# ──────────────────────────────────────────────

class DocumentoContableService:

    @staticmethod
    def list_documentos(proyecto: Project, fase_id: str | None = None) -> QuerySet:
        """
        Retorna documentos contables del proyecto.
        Opcionalmente filtra por fase.
        """
        qs = (
            AccountingDocument.all_objects
            .filter(proyecto=proyecto)
            .select_related('fase')
            .order_by('-fecha_documento')
        )
        if fase_id:
            qs = qs.filter(fase_id=fase_id)
        return qs

    @staticmethod
    def get_documento(documento_id: str) -> AccountingDocument:
        return (
            AccountingDocument.objects
            .select_related('proyecto', 'fase')
            .get(id=documento_id)
        )


# ──────────────────────────────────────────────
# HitoService
# ──────────────────────────────────────────────

class HitoService:

    @staticmethod
    def list_hitos(proyecto: Project) -> QuerySet:
        """Retorna hitos del proyecto ordenados por fecha de creación."""
        return (
            Milestone.all_objects
            .filter(proyecto=proyecto)
            .select_related('fase', 'documento_factura')
            .order_by('created_at')
        )

    @staticmethod
    @transaction.atomic
    def create_hito(proyecto: Project, data: dict) -> Milestone:
        """
        Crea un hito facturable validando que el porcentaje total
        de hitos no supere el 100% del proyecto.
        """
        from django.db.models import Sum as DbSum

        porcentaje_nuevo = data.get('porcentaje_proyecto', Decimal('0'))

        # Bloquear proyecto para evitar race conditions
        proyecto_locked = Project.all_objects.select_for_update().get(id=proyecto.id)

        porcentaje_existente = (
            Milestone.all_objects
            .filter(proyecto=proyecto_locked)
            .aggregate(total=DbSum('porcentaje_proyecto'))
            .get('total') or Decimal('0')
        )

        if porcentaje_existente + porcentaje_nuevo > Decimal('100'):
            disponible = Decimal('100') - porcentaje_existente
            raise ValidationError(
                f'El porcentaje total de hitos superaría el 100%. '
                f'Disponible: {disponible}%. Solicitado: {porcentaje_nuevo}%.'
            )

        hito = Milestone(
            proyecto=proyecto_locked,
            company=proyecto_locked.company,
            **data,
        )
        hito.full_clean()
        hito.save()

        logger.info(
            'Milestone creado',
            extra={
                'hito_id': str(hito.id),
                'proyecto_id': str(proyecto.id),
                'porcentaje': str(porcentaje_nuevo),
            },
        )
        return hito

    @staticmethod
    @transaction.atomic
    def generar_factura(hito: Milestone, user) -> Milestone:
        """
        Marca el hito como facturado y registra la solicitud.
        El agente Go procesará la solicitud y creará la factura en Saiopen.
        """
        if not hito.facturable:
            raise ValidationError('Este hito no está marcado como facturable.')
        if hito.facturado:
            raise ValidationError('Este hito ya fue facturado.')
        if not hito.proyecto.sincronizado_con_saiopen:
            raise ValidationError(
                'El proyecto debe estar sincronizado con Saiopen para generar facturas.'
            )

        hito.facturado         = True
        hito.fecha_facturacion = timezone.now().date()
        hito.save(update_fields=['facturado', 'fecha_facturacion', 'updated_at'])

        logger.info(
            'Solicitud de factura generada para hito',
            extra={
                'hito_id': str(hito.id),
                'proyecto_id': str(hito.proyecto_id),
                'valor_facturar': str(hito.valor_facturar),
                'user': user.email,
            },
        )
        return hito


# ──────────────────────────────────────────────
# ActividadService — catálogo global
# ──────────────────────────────────────────────

class ActividadService:

    @staticmethod
    def list_actividades(company=None) -> QuerySet:
        qs = Activity.all_objects.filter(activo=True)
        if company is not None:
            qs = qs.filter(company=company)
        return qs

    @staticmethod
    def get_actividad(pk) -> Activity:
        return Activity.all_objects.get(id=pk)

    @staticmethod
    @transaction.atomic
    def create_actividad(data: dict, user) -> Activity:
        company = user.company

        # Auto-generar código: usar consecutivo seleccionado por el usuario, fallback a secuencial
        consecutivo_id = data.pop('consecutivo_id', None)
        if not data.get('codigo'):
            from apps.core.services import generar_consecutivo
            codigo = generar_consecutivo(str(consecutivo_id)) if consecutivo_id else None
            if not codigo:
                count = Activity.all_objects.filter(company=company).count()
                codigo = f'ACT-{str(count + 1).zfill(3)}'
            data = {**data, 'codigo': codigo}

        actividad = Activity(company=company, **data)
        actividad.full_clean()
        actividad.save()

        logger.info(
            'Actividad creada',
            extra={'actividad_id': str(actividad.id), 'codigo': actividad.codigo},
        )
        return actividad

    @staticmethod
    def update_actividad(actividad: Activity, data: dict) -> Activity:
        for key, value in data.items():
            setattr(actividad, key, value)
        actividad.full_clean()
        actividad.save()
        logger.info('Actividad actualizada', extra={'actividad_id': str(actividad.id)})
        return actividad

    @staticmethod
    def soft_delete_actividad(actividad: Activity) -> None:
        # Verificar que no tenga asignaciones activas antes de eliminar
        asignaciones = ProjectActivity.all_objects.filter(actividad=actividad).count()
        if asignaciones > 0:
            raise ValidationError(
                f'No se puede eliminar la actividad "{actividad.codigo}" porque está '
                f'asignada a {asignaciones} proyecto(s).'
            )
        actividad.activo = False
        actividad.save(update_fields=['activo', 'updated_at'])
        logger.info('Actividad eliminada (soft)', extra={'actividad_id': str(actividad.id)})


# ──────────────────────────────────────────────
# ActividadProyectoService — asignación al proyecto
# ──────────────────────────────────────────────

class ActividadProyectoService:

    @staticmethod
    def list_actividades_proyecto(proyecto: Project, fase_id: str | None = None) -> QuerySet:
        qs = (
            ProjectActivity.all_objects
            .filter(proyecto=proyecto)
            .select_related('actividad', 'fase')
        )
        if fase_id:
            qs = qs.filter(fase_id=fase_id)
        return qs

    @staticmethod
    @transaction.atomic
    def asignar_actividad(proyecto: Project, data: dict) -> ProjectActivity:
        # Si no se especifica costo_unitario, usar el base del catálogo
        actividad: Actividad = data['actividad']
        if not data.get('costo_unitario'):
            data = {**data, 'costo_unitario': actividad.costo_unitario_base}

        ap = ProjectActivity(
            company=proyecto.company,
            proyecto=proyecto,
            **data,
        )
        ap.full_clean()
        ap.save()

        logger.info(
            'Actividad asignada al proyecto',
            extra={
                'proyecto_id': str(proyecto.id),
                'actividad_id': str(actividad.id),
                'actividad_codigo': actividad.codigo,
            },
        )
        return ap

    @staticmethod
    def update_actividad_proyecto(ap: ProjectActivity, data: dict) -> ProjectActivity:
        for key, value in data.items():
            setattr(ap, key, value)
        ap.full_clean()
        ap.save()
        logger.info('ActividadProyecto actualizada', extra={'id': str(ap.id)})
        return ap

    @staticmethod
    def desasignar_actividad(ap: ProjectActivity) -> None:
        estados_bloqueados = [ProjectStatus.PLANNED, ProjectStatus.IN_PROGRESS]
        if ap.proyecto.estado in estados_bloqueados:
            raise ValidationError(
                f'No se pueden eliminar actividades de un proyecto en estado '
                f'"{ap.proyecto.get_estado_display()}". Solo se permite en estado Borrador.'
            )
        logger.info(
            'Actividad desasignada del proyecto',
            extra={'id': str(ap.id), 'proyecto_id': str(ap.proyecto_id)},
        )
        ap.delete()


# ──────────────────────────────────────────────
# Feature #8 — PDF Export
# ──────────────────────────────────────────────

class ProyectoExportService:
    """
    Genera reportes PDF de proyectos usando WeasyPrint.
    """

    @staticmethod
    def generate_pdf(proyecto: Project, include_gantt: bool = True, include_budget: bool = True) -> bytes:
        """
        Genera el PDF del proyecto y retorna los bytes del PDF.
        """
        from django.template.loader import render_to_string

        if not _WEASYPRINT_AVAILABLE:
            raise ImportError(
                'WeasyPrint no está instalado. Ejecutar: pip install weasyprint==60.1'
            )

        context = ProyectoExportService._prepare_context(
            proyecto, include_gantt=include_gantt, include_budget=include_budget
        )
        html_string = render_to_string('proyectos/export/pdf_report.html', context)
        css_string  = ProyectoExportService._get_pdf_css()

        pdf_bytes = WeasyHTML(string=html_string).write_pdf(
            stylesheets=[WeasyCSS(string=css_string)]
        )
        logger.info(
            'PDF de proyecto generado',
            extra={
                'proyecto_id': str(proyecto.id),
                'codigo': proyecto.codigo,
                'include_gantt': include_gantt,
                'include_budget': include_budget,
            },
        )
        return pdf_bytes

    @staticmethod
    def _prepare_context(proyecto: Project, include_gantt: bool = True, include_budget: bool = True) -> dict:
        """Recopila todos los datos necesarios para el template del PDF."""
        import pytz
        from django.utils import timezone
        from django.db.models import Sum, Count

        fases = list(
            proyecto.phases
            .filter(activo=True)
            .order_by('orden')
            .prefetch_related('tasks')
        )

        # KPIs básicos
        total_tareas = proyecto.tasks.count()
        tareas_completadas = proyecto.tasks.filter(estado='completed').count()
        tareas_en_progreso = proyecto.tasks.filter(estado='in_progress').count()
        tareas_pendientes  = proyecto.tasks.filter(estado='todo').count()

        pct_tareas_completadas = (
            round(tareas_completadas / total_tareas * 100, 1) if total_tareas > 0 else 0
        )

        # Resumen presupuestario (desde fases activas)
        fases_agg = proyecto.phases.filter(activo=True).aggregate(
            total_mano_obra=Sum('presupuesto_mano_obra'),
            total_materiales=Sum('presupuesto_materiales'),
            total_subcontratos=Sum('presupuesto_subcontratos'),
            total_equipos=Sum('presupuesto_equipos'),
            total_otros=Sum('presupuesto_otros'),
        )

        def _d(val):
            return val or Decimal('0')

        presupuesto_costos = (
            _d(fases_agg['total_mano_obra'])
            + _d(fases_agg['total_materiales'])
            + _d(fases_agg['total_subcontratos'])
            + _d(fases_agg['total_equipos'])
            + _d(fases_agg['total_otros'])
        )

        aiu_factor = (
            proyecto.porcentaje_administracion
            + proyecto.porcentaje_imprevistos
            + proyecto.porcentaje_utilidad
        ) / Decimal('100')
        precio_venta_aiu = presupuesto_costos * (1 + aiu_factor)

        docs_agg = proyecto.documents.aggregate(ejecutado=Sum('valor_neto'))
        costo_ejecutado = _d(docs_agg['ejecutado'])

        budget_summary = {
            'presupuesto_total':    str(proyecto.presupuesto_total),
            'presupuesto_costos':   str(presupuesto_costos),
            'precio_venta_aiu':     str(precio_venta_aiu.quantize(Decimal('0.01'))),
            'costo_ejecutado':      str(costo_ejecutado),
            'desglose': {
                'mano_obra':    str(_d(fases_agg['total_mano_obra'])),
                'materiales':   str(_d(fases_agg['total_materiales'])),
                'subcontratos': str(_d(fases_agg['total_subcontratos'])),
                'equipos':      str(_d(fases_agg['total_equipos'])),
                'otros':        str(_d(fases_agg['total_otros'])),
            },
            'aiu': {
                'administracion': str(proyecto.porcentaje_administracion),
                'imprevistos':    str(proyecto.porcentaje_imprevistos),
                'utilidad':       str(proyecto.porcentaje_utilidad),
                'valor_aiu':      str((precio_venta_aiu - presupuesto_costos).quantize(Decimal('0.01'))),
            },
        }

        logo_base64 = ProyectoExportService._get_tenant_logo_base64(proyecto)

        return {
            'proyecto':               proyecto,
            'fases':                  fases,
            'include_gantt':          include_gantt,
            'include_budget':         include_budget,
            'budget_summary':         budget_summary,
            'logo_base64':            logo_base64,
            'fecha_generacion':       timezone.now().astimezone(pytz.timezone('America/Bogota')).strftime('%d/%m/%Y %I:%M %p'),
            'kpis': {
                'total_tareas':             total_tareas,
                'tareas_completadas':       tareas_completadas,
                'tareas_en_progreso':       tareas_en_progreso,
                'tareas_pendientes':        tareas_pendientes,
                'pct_tareas_completadas':   pct_tareas_completadas,
                'porcentaje_avance_fisico': str(proyecto.porcentaje_avance),
            },
        }

    @staticmethod
    def _get_pdf_css() -> str:
        """Retorna el string CSS para el PDF (A4, colores corporativos)."""
        return """
        @page {
            size: A4;
            margin: 2cm 2cm 2.5cm 2cm;
            @bottom-right {
                content: "Página " counter(page) " de " counter(pages);
                font-size: 9pt;
                color: #666;
            }
        }

        * { box-sizing: border-box; }

        body {
            font-family: 'Liberation Sans', Arial, sans-serif;
            font-size: 10pt;
            color: #212121;
            line-height: 1.5;
        }

        h1 { font-size: 22pt; font-weight: 800; color: #1a237e; margin: 0 0 8px 0; }
        h2 { font-size: 14pt; font-weight: 700; color: #1a237e; border-bottom: 2px solid #1a237e; padding-bottom: 4px; margin-top: 20px; }
        h3 { font-size: 11pt; font-weight: 700; color: #283593; margin-top: 16px; }

        .portada {
            page-break-after: always;
            padding: 60px 0 40px 0;
        }

        .portada .logo { max-height: 80px; margin-bottom: 40px; }
        .portada .titulo { font-family: 'Liberation Sans', Arial, sans-serif; font-size: 28pt; font-weight: 800; color: #1a237e; margin-bottom: 8px; letter-spacing: -0.5pt; }
        .portada .subtitulo { font-size: 14pt; color: #3949ab; margin-bottom: 4px; }
        .portada .meta { color: #666; font-size: 9pt; margin-top: 40px; }

        .badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 8pt;
            font-weight: bold;
        }
        .badge-draft       { background: #e3f2fd; color: #0d47a1; }
        .badge-planned     { background: #e8f5e9; color: #1b5e20; }
        .badge-in_progress { background: #fff3e0; color: #e65100; }
        .badge-suspended   { background: #fce4ec; color: #880e4f; }
        .badge-closed      { background: #ede7f6; color: #4a148c; }
        .badge-cancelled   { background: #f5f5f5; color: #616161; }

        table {
            width: 100%;
            border-collapse: collapse;
            margin: 10px 0;
            font-size: 9pt;
        }
        th {
            background: #1a237e;
            color: white;
            padding: 6px 8px;
            text-align: left;
        }
        td { padding: 5px 8px; border-bottom: 1px solid #e0e0e0; }
        tr:nth-child(even) td { background: #f5f7ff; }

        .kpi-grid {
            display: table;
            width: 100%;
            margin: 12px 0;
        }
        .kpi-card {
            display: table-cell;
            width: 25%;
            background: #f5f7ff;
            border: 1px solid #c5cae9;
            border-radius: 6px;
            padding: 12px;
            text-align: center;
            vertical-align: middle;
        }
        .kpi-value { font-size: 20pt; font-weight: bold; color: #1a237e; }
        .kpi-label { font-size: 8pt; color: #666; }

        .progress-bar-container {
            width: 100%;
            background: #e0e0e0;
            border-radius: 4px;
            height: 10px;
            margin: 4px 0;
        }
        .progress-bar {
            height: 10px;
            background: #3f51b5;
            border-radius: 4px;
        }

        .fase-block {
            border: 1px solid #c5cae9;
            border-radius: 6px;
            padding: 12px;
            margin: 12px 0;
            page-break-inside: avoid;
        }

        .gantt-placeholder {
            background: #f5f7ff;
            border: 2px dashed #c5cae9;
            border-radius: 6px;
            padding: 20px;
            text-align: center;
            color: #9e9e9e;
            font-style: italic;
        }

        .footer-note {
            font-size: 8pt;
            color: #999;
            text-align: center;
            margin-top: 20px;
        }
        """

    @staticmethod
    def _get_tenant_logo_base64(proyecto: Project):
        """
        Intenta obtener el logo de la empresa como base64.
        Retorna None si no existe o hay algún error.
        """
        try:
            company = proyecto.company
            logo_field = getattr(company, 'logo', None)
            if not logo_field:
                return None
            logo_url = str(logo_field)
            if not logo_url:
                return None

            import base64
            import urllib.request

            # Si es URL HTTP
            if logo_url.startswith('http'):
                with urllib.request.urlopen(logo_url, timeout=5) as response:
                    data = response.read()
                    return base64.b64encode(data).decode('utf-8')

            # Si es path local
            import os
            from django.conf import settings as django_settings
            full_path = os.path.join(django_settings.MEDIA_ROOT, logo_url.lstrip('/'))
            if os.path.exists(full_path):
                with open(full_path, 'rb') as f:
                    data = f.read()
                    return base64.b64encode(data).decode('utf-8')

        except Exception as exc:
            logger.warning(
                'No se pudo cargar el logo de la empresa',
                extra={'proyecto_id': str(proyecto.id), 'error': str(exc)},
            )
        return None


# ──────────────────────────────────────────────
# Feature #8 — Project Templates
# ──────────────────────────────────────────────

class PlantillaProyectoService:
    """
    Gestiona las plantillas de proyecto y la creación de proyectos desde ellas.
    """

    @staticmethod
    def list_plantillas(company, is_active: bool = True):
        """Retorna plantillas activas de la empresa con conteos de fases y tareas."""
        from apps.proyectos.models import PlantillaProyecto

        return (
            PlantillaProyecto.objects
            .filter(company=company, is_active=is_active)
            .prefetch_related('fases_plantilla__tareas_plantilla')
            .order_by('tipo', 'nombre')
        )

    @staticmethod
    def get_plantilla(plantilla_id: str, company):
        """Retorna una plantilla de la empresa con todas sus relaciones cargadas."""
        from apps.proyectos.models import PlantillaProyecto
        return (
            PlantillaProyecto.objects
            .prefetch_related(
                'fases_plantilla',
                'fases_plantilla__tareas_plantilla',
                'fases_plantilla__tareas_plantilla__actividad_saiopen',
                'fases_plantilla__tareas_plantilla__sucesoras_plantilla',
            )
            .get(id=plantilla_id, company=company, is_active=True)
        )

    @staticmethod
    @transaction.atomic
    def create_plantilla(data: dict, company) -> 'PlantillaProyecto':
        """Crea una plantilla con fases y tareas anidadas."""
        from apps.proyectos.models import PlantillaProyecto, PlantillaFase, PlantillaTarea

        plantilla = PlantillaProyecto(
            company=company,
            nombre=data['nombre'],
            descripcion=data.get('descripcion', ''),
            tipo=data['tipo'],
            icono=data.get('icono', 'folder'),
            duracion_estimada=data.get('duracion_estimada', 30),
            is_active=True,
        )
        plantilla.save()

        for fase_data in data.get('fases', []):
            fase = PlantillaFase(
                company=company,
                plantilla_proyecto=plantilla,
                nombre=fase_data['nombre'],
                descripcion=fase_data.get('descripcion', ''),
                orden=fase_data.get('orden', 0),
                porcentaje_duracion=fase_data.get('porcentaje_duracion', Decimal('100.00')),
            )
            fase.save()

            for tarea_data in fase_data.get('tareas', []):
                tarea = PlantillaTarea(
                    company=company,
                    plantilla_fase=fase,
                    nombre=tarea_data['nombre'],
                    descripcion=tarea_data.get('descripcion', ''),
                    orden=tarea_data.get('orden', 0),
                    duracion_dias=tarea_data.get('duracion_dias', 1),
                    actividad_saiopen_id=tarea_data.get('actividad_saiopen_id'),
                )
                tarea.save()

        logger.info(
            'Plantilla de proyecto creada',
            extra={
                'plantilla_id': str(plantilla.id),
                'nombre': plantilla.nombre,
                'company': str(company.id),
            },
        )
        return plantilla

    @staticmethod
    @transaction.atomic
    def update_plantilla(plantilla_id: str, data: dict, company) -> 'PlantillaProyecto':
        """Actualiza una plantilla, reemplazando fases y tareas."""
        from apps.proyectos.models import PlantillaProyecto, PlantillaFase, PlantillaTarea

        try:
            plantilla = PlantillaProyecto.objects.get(id=plantilla_id, company=company)
        except PlantillaProyecto.DoesNotExist:
            raise ProyectoException({'plantilla_id': 'Plantilla no encontrada o no pertenece a su empresa.'})

        plantilla.nombre           = data.get('nombre', plantilla.nombre)
        plantilla.descripcion      = data.get('descripcion', plantilla.descripcion)
        plantilla.tipo             = data.get('tipo', plantilla.tipo)
        plantilla.icono            = data.get('icono', plantilla.icono)
        plantilla.duracion_estimada = data.get('duracion_estimada', plantilla.duracion_estimada)
        plantilla.save()

        # Reemplazar fases y tareas si se envían
        if 'fases' in data:
            plantilla.fases_plantilla.all().delete()
            for fase_data in data['fases']:
                fase = PlantillaFase(
                    company=company,
                    plantilla_proyecto=plantilla,
                    nombre=fase_data['nombre'],
                    descripcion=fase_data.get('descripcion', ''),
                    orden=fase_data.get('orden', 0),
                    porcentaje_duracion=fase_data.get('porcentaje_duracion', Decimal('100.00')),
                )
                fase.save()

                for tarea_data in fase_data.get('tareas', []):
                    tarea = PlantillaTarea(
                        company=company,
                        plantilla_fase=fase,
                        nombre=tarea_data['nombre'],
                        descripcion=tarea_data.get('descripcion', ''),
                        orden=tarea_data.get('orden', 0),
                        duracion_dias=tarea_data.get('duracion_dias', 1),
                        actividad_saiopen_id=tarea_data.get('actividad_saiopen_id'),
                    )
                    tarea.save()

        logger.info(
            'Plantilla de proyecto actualizada',
            extra={
                'plantilla_id': str(plantilla.id),
                'nombre': plantilla.nombre,
                'company': str(company.id),
            },
        )
        return plantilla

    @staticmethod
    def delete_plantilla(plantilla_id: str, company) -> None:
        """Elimina una plantilla (solo si pertenece a la empresa)."""
        from apps.proyectos.models import PlantillaProyecto

        try:
            plantilla = PlantillaProyecto.objects.get(id=plantilla_id, company=company)
        except PlantillaProyecto.DoesNotExist:
            raise ProyectoException({'plantilla_id': 'Plantilla no encontrada o no pertenece a su empresa.'})

        nombre = plantilla.nombre
        plantilla.delete()

        logger.info(
            'Plantilla de proyecto eliminada',
            extra={
                'plantilla_id': plantilla_id,
                'nombre': nombre,
                'company': str(company.id),
            },
        )

    @staticmethod
    @transaction.atomic
    def create_from_template(
        plantilla_id: str,
        nombre: str,
        descripcion: str,
        planned_start,
        user,
        cliente_id: str | None = None,
    ) -> Project:
        """
        Clona una plantilla a un proyecto real con fases, tareas y dependencias.

        Algoritmo de fechas:
        - Cada fase ocupa (porcentaje_duracion / 100) * duracion_estimada días.
        - Las fases son secuenciales: cada una empieza el día después de que termina la anterior.
        - Las tareas se distribuyen dentro de la fase: fecha_inicio = fase.start + (orden-1)*duracion_dias.
        """
        from datetime import timedelta
        from apps.proyectos.models import PlantillaProyecto, PlantillaDependencia, TaskDependency

        company = getattr(user, 'effective_company', None) or user.company

        try:
            plantilla = PlantillaProyecto.all_objects.prefetch_related(
                'fases_plantilla__tareas_plantilla__actividad_saiopen'
            ).get(id=plantilla_id, company=company, is_active=True)
        except PlantillaProyecto.DoesNotExist:
            raise ProyectoException({'template_id': 'Plantilla no encontrada o inactiva.'})

        # Crear el proyecto base
        codigo = ProyectoService._generar_codigo(company)
        proyecto = Project(
            company=company,
            codigo=codigo,
            nombre=nombre,
            tipo=plantilla.tipo,
            estado=ProjectStatus.DRAFT,
            cliente_id=cliente_id or '',
            cliente_nombre='',
            gerente=user,
            fecha_inicio_planificada=planned_start,
            fecha_fin_planificada=planned_start + timedelta(days=max(plantilla.duracion_estimada, 1)),
            presupuesto_total=Decimal('0'),
        )
        proyecto.save()
        logger.info(
            'Proyecto creado desde plantilla',
            extra={
                'proyecto_id': str(proyecto.id),
                'plantilla_id': str(plantilla.id),
                'plantilla_nombre': plantilla.nombre,
            },
        )

        # Crear fases con fechas calculadas
        from apps.proyectos.models import ProjectActivity

        fase_map = {}        # plantilla_fase.id → Phase
        tarea_map = {}       # plantilla_tarea.id → Task
        # (actividad_id, fase_id) → acumula cantidad para evitar duplicados en unique_together
        activities_to_create: dict[tuple, dict] = {}

        current_date = planned_start
        fases_plantilla = list(plantilla.fases_plantilla.order_by('orden'))

        for idx, fase_tmpl in enumerate(fases_plantilla):
            duracion_fase = max(
                round(float(fase_tmpl.porcentaje_duracion) / 100 * plantilla.duracion_estimada),
                1,
            )
            fecha_inicio_fase = current_date
            fecha_fin_fase    = current_date + timedelta(days=duracion_fase - 1)

            fase = Phase(
                company=company,
                proyecto=proyecto,
                nombre=fase_tmpl.nombre,
                descripcion=fase_tmpl.descripcion,
                orden=fase_tmpl.orden,
                fecha_inicio_planificada=fecha_inicio_fase,
                fecha_fin_planificada=fecha_fin_fase,
            )
            fase.save()
            fase_map[str(fase_tmpl.id)] = fase

            logger.info(
                'Fase creada desde plantilla',
                extra={
                    'fase_id': str(fase.id),
                    'fase_plantilla_id': str(fase_tmpl.id),
                    'proyecto_id': str(proyecto.id),
                },
            )

            # Crear tareas de la fase
            tareas_plantilla = list(fase_tmpl.tareas_plantilla.order_by('orden'))
            for tarea_tmpl in tareas_plantilla:
                fecha_inicio_tarea = fecha_inicio_fase + timedelta(
                    days=(tarea_tmpl.orden - 1) * tarea_tmpl.duracion_dias
                )
                fecha_fin_tarea = fecha_inicio_tarea + timedelta(days=tarea_tmpl.duracion_dias - 1)

                tarea = Task(
                    company=company,
                    fase=fase,
                    proyecto=proyecto,
                    nombre=tarea_tmpl.nombre,
                    descripcion=tarea_tmpl.descripcion,
                    prioridad=tarea_tmpl.prioridad,
                    fecha_inicio=fecha_inicio_tarea,
                    fecha_fin=fecha_fin_tarea,
                    horas_estimadas=Decimal(str(tarea_tmpl.duracion_dias * 8)),
                )
                tarea.save()
                tarea_map[str(tarea_tmpl.id)] = tarea

                # Recolectar actividades para crear ProjectActivity
                if tarea_tmpl.actividad_saiopen_id and tarea_tmpl.actividad_saiopen:
                    act = tarea_tmpl.actividad_saiopen
                    key = (str(act.id), str(fase.id))
                    cantidad = Decimal(str(tarea_tmpl.duracion_dias))
                    if key in activities_to_create:
                        activities_to_create[key]['cantidad_planificada'] += cantidad
                    else:
                        activities_to_create[key] = {
                            'company': company,
                            'proyecto': proyecto,
                            'actividad': act,
                            'fase': fase,
                            'cantidad_planificada': cantidad,
                            'costo_unitario': act.costo_unitario_base,
                        }

            # Avanzar al siguiente día después de esta fase
            current_date = fecha_fin_fase + timedelta(days=1)

        # Crear ProjectActivity para todas las actividades recolectadas
        for data in activities_to_create.values():
            ProjectActivity.objects.create(**data)
            logger.info(
                'ProjectActivity creada desde plantilla',
                extra={
                    'proyecto_id': str(proyecto.id),
                    'actividad_id': str(data['actividad'].id),
                    'actividad_codigo': data['actividad'].codigo,
                    'cantidad': str(data['cantidad_planificada']),
                },
            )

        # Actualizar fecha fin real del proyecto
        if fecha_fin_fase:
            Project.objects.filter(id=proyecto.id).update(fecha_fin_planificada=fecha_fin_fase)

        # Clonar dependencias entre tareas
        dependencias = PlantillaDependencia.objects.filter(
            tarea_predecesora__plantilla_fase__plantilla_proyecto=plantilla
        ).select_related('tarea_predecesora', 'tarea_sucesora')

        for dep_tmpl in dependencias:
            pred_key = str(dep_tmpl.tarea_predecesora_id)
            succ_key = str(dep_tmpl.tarea_sucesora_id)

            if pred_key not in tarea_map or succ_key not in tarea_map:
                logger.warning(
                    'Dependencia de plantilla sin tareas correspondientes',
                    extra={
                        'dep_id': str(dep_tmpl.id),
                        'pred': pred_key,
                        'succ': succ_key,
                    },
                )
                continue

            TaskDependency.objects.create(
                company=company,
                tarea_predecesora=tarea_map[pred_key],
                tarea_sucesora=tarea_map[succ_key],
                tipo_dependencia=dep_tmpl.tipo_dependencia,
                retraso_dias=dep_tmpl.lag_time,
            )

        logger.info(
            'Proyecto creado exitosamente desde plantilla',
            extra={
                'proyecto_id': str(proyecto.id),
                'codigo': proyecto.codigo,
                'fases_creadas': len(fase_map),
                'tareas_creadas': len(tarea_map),
            },
        )
        return proyecto


# ──────────────────────────────────────────────
# Feature #8 — Excel Import
# ──────────────────────────────────────────────

class ProyectoImportService:
    """
    Importa proyectos desde archivos Excel (.xlsx/.xls).

    Estructura esperada del archivo:
    - Hoja 'Datos Proyecto': columnas nombre, tipo, cliente_id, cliente_nombre,
                              gerente_email, fecha_inicio_planificada, fecha_fin_planificada,
                              presupuesto_total (opcional), descripcion (opcional)
    - Hoja 'Fases': columnas nombre, orden, fecha_inicio, fecha_fin,
                    descripcion (opcional), presupuesto_mano_obra (opcional),
                    presupuesto_materiales (opcional), presupuesto_subcontratos (opcional),
                    presupuesto_equipos (opcional), presupuesto_otros (opcional)
    - Hoja 'Tareas': columnas nombre, fase_orden, descripcion (opcional),
                     fecha_inicio (opcional), fecha_fin (opcional),
                     prioridad (1-4, opcional), horas_estimadas (opcional)
    - Hoja 'Dependencias' (opcional): pred_nombre, succ_nombre, tipo (FS/SS/FF), lag (opcional)
    """

    @staticmethod
    @transaction.atomic
    def import_from_excel(file, company, user) -> dict:
        """
        Lee el archivo Excel e importa el proyecto.
        Retorna {'success': bool, 'proyecto': Proyecto|None, 'errors': [], 'stats': {}}.
        """
        try:
            import openpyxl
        except ImportError:
            return {
                'success': False,
                'proyecto': None,
                'errors': ['openpyxl no está instalado.'],
                'stats': {},
            }

        errors = []
        stats  = {'fases': 0, 'tareas': 0, 'dependencias': 0}

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            return {
                'success': False,
                'proyecto': None,
                'errors': [f'No se pudo leer el archivo Excel: {exc}'],
                'stats': stats,
            }

        try:
            with transaction.atomic():
                # Leer datos del proyecto
                proyecto_data = ProyectoImportService._read_proyecto_data(wb, company, user)

                # Crear proyecto
                proyecto = ProyectoService.create_proyecto(proyecto_data, user)

                # Leer y crear fases
                fases_data = ProyectoImportService._read_fases_data(wb)
                fase_map   = {}  # orden → Phase

                for fase_d in fases_data:
                    fase = Phase(
                        company=company,
                        proyecto=proyecto,
                        nombre=fase_d['nombre'],
                        descripcion=fase_d.get('descripcion', ''),
                        orden=fase_d['orden'],
                        fecha_inicio_planificada=fase_d['fecha_inicio'],
                        fecha_fin_planificada=fase_d['fecha_fin'],
                        presupuesto_mano_obra=Decimal(str(fase_d.get('presupuesto_mano_obra', 0) or 0)),
                        presupuesto_materiales=Decimal(str(fase_d.get('presupuesto_materiales', 0) or 0)),
                        presupuesto_subcontratos=Decimal(str(fase_d.get('presupuesto_subcontratos', 0) or 0)),
                        presupuesto_equipos=Decimal(str(fase_d.get('presupuesto_equipos', 0) or 0)),
                        presupuesto_otros=Decimal(str(fase_d.get('presupuesto_otros', 0) or 0)),
                    )
                    fase.save()
                    fase_map[fase_d['orden']] = fase
                    stats['fases'] += 1

                # Leer y crear tareas
                tareas_data = ProyectoImportService._read_tareas_data(wb)
                tarea_map   = {}  # nombre → Task

                for tarea_d in tareas_data:
                    fase_orden = tarea_d.get('fase_orden')
                    fase       = fase_map.get(fase_orden)

                    if not fase:
                        errors.append(
                            f"Tarea '{tarea_d.get('nombre')}': fase con orden {fase_orden} no encontrada."
                        )
                        continue

                    tarea = Task(
                        company=company,
                        fase=fase,
                        proyecto=proyecto,
                        nombre=tarea_d['nombre'],
                        descripcion=tarea_d.get('descripcion', ''),
                        prioridad=int(tarea_d.get('prioridad', 2) or 2),
                        horas_estimadas=Decimal(str(tarea_d.get('horas_estimadas', 0) or 0)),
                        fecha_inicio=tarea_d.get('fecha_inicio'),
                        fecha_fin=tarea_d.get('fecha_fin'),
                    )
                    tarea.save()
                    tarea_map[tarea_d['nombre']] = tarea
                    stats['tareas'] += 1

                # Leer y crear dependencias (opcional)
                deps_data = ProyectoImportService._read_dependencias_data(wb)
                for dep_d in deps_data:
                    pred = tarea_map.get(dep_d.get('pred_nombre'))
                    succ = tarea_map.get(dep_d.get('succ_nombre'))

                    if not pred or not succ:
                        errors.append(
                            f"Dependencia '{dep_d.get('pred_nombre')}' → '{dep_d.get('succ_nombre')}': "
                            f"tarea no encontrada (se omite)."
                        )
                        continue

                    try:
                        from apps.proyectos.models import TaskDependency
                        tipo_str = str(dep_d.get('tipo', 'FS')).upper()
                        tipo = tipo_str if tipo_str in ('FS', 'SS', 'FF') else 'FS'

                        TaskDependency.objects.create(
                            company=company,
                            tarea_predecesora=pred,
                            tarea_sucesora=succ,
                            tipo_dependencia=tipo,
                            retraso_dias=int(dep_d.get('lag', 0) or 0),
                        )
                        stats['dependencias'] += 1
                    except Exception as dep_exc:
                        errors.append(
                            f"Dependencia '{dep_d.get('pred_nombre')}' → '{dep_d.get('succ_nombre')}': "
                            f"{dep_exc} (se omite)."
                        )

                logger.info(
                    'Proyecto importado desde Excel',
                    extra={
                        'proyecto_id': str(proyecto.id),
                        'codigo': proyecto.codigo,
                        'fases': stats['fases'],
                        'tareas': stats['tareas'],
                        'dependencias': stats['dependencias'],
                        'errores': len(errors),
                    },
                )

            return {
                'success': True,
                'proyecto': proyecto,
                'errors': errors,
                'stats': stats,
            }

        except Exception as exc:
            logger.error(
                'Error importando proyecto desde Excel',
                extra={'error': str(exc), 'user': user.email},
            )
            return {
                'success': False,
                'proyecto': None,
                'errors': [str(exc)],
                'stats': stats,
            }

    @staticmethod
    def _read_proyecto_data(wb, company, user) -> dict:
        """Lee datos básicos del proyecto desde la hoja 'Datos Proyecto'."""
        sheet_name = 'Datos Proyecto'
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("La hoja 'Datos Proyecto' no tiene datos.")

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        values  = rows[1]
        data    = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}

        nombre = data.get('nombre')
        if not nombre:
            raise ValueError("Campo obligatorio 'nombre' no encontrado en 'Datos Proyecto'.")

        tipo_val = str(data.get('tipo', 'other')).lower()
        valid_tipos = [t[0] for t in ProjectType.choices]
        tipo = tipo_val if tipo_val in valid_tipos else 'other'

        fecha_inicio = ProyectoImportService._parse_date(data.get('fecha_inicio_planificada'))
        fecha_fin    = ProyectoImportService._parse_date(data.get('fecha_fin_planificada'))

        if not fecha_inicio:
            raise ValueError("Campo obligatorio 'fecha_inicio_planificada' no encontrado o inválido.")
        if not fecha_fin:
            raise ValueError("Campo obligatorio 'fecha_fin_planificada' no encontrado o inválido.")

        # Resolver gerente — intentar por email primero, luego usar el user que importa
        gerente = user
        gerente_email = data.get('gerente_email')
        if gerente_email:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                gerente = User.objects.get(email=str(gerente_email).strip(), company=company)
            except User.DoesNotExist:
                pass  # Usar el user que importa como fallback

        presupuesto = data.get('presupuesto_total', 0)
        try:
            presupuesto = Decimal(str(presupuesto)) if presupuesto else Decimal('0')
        except Exception:
            presupuesto = Decimal('0')

        # cliente_nit → se guarda en cliente_id (referencia Saiopen).
        # Si el usuario escribe el NIT/cédula lo usamos directamente.
        # Si no viene, queda vacío (blank=True en el modelo).
        cliente_nit = str(data.get('cliente_nit', '') or '').strip()
        cliente_nombre = str(data.get('cliente_nombre', '') or '').strip()

        return {
            'nombre': str(nombre).strip(),
            'tipo': tipo,
            'cliente_id': cliente_nit,
            'cliente_nombre': cliente_nombre,
            'gerente': gerente.id,
            'fecha_inicio_planificada': fecha_inicio,
            'fecha_fin_planificada': fecha_fin,
            'presupuesto_total': presupuesto,
        }

    @staticmethod
    def _read_fases_data(wb) -> list:
        """Lee fases desde la hoja 'Fases'."""
        sheet_name = 'Fases'
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo.")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError("La hoja 'Fases' no tiene datos.")

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        fases   = []

        for i, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}

            nombre = data.get('nombre')
            if not nombre:
                raise ValueError(f"Fila {i} en 'Fases': campo 'nombre' obligatorio.")

            orden = data.get('orden')
            try:
                orden = int(orden) if orden is not None else i - 1
            except (ValueError, TypeError):
                orden = i - 1

            fecha_inicio = ProyectoImportService._parse_date(data.get('fecha_inicio'))
            fecha_fin    = ProyectoImportService._parse_date(data.get('fecha_fin'))

            if not fecha_inicio:
                raise ValueError(f"Fila {i} en 'Fases': campo 'fecha_inicio' obligatorio.")
            if not fecha_fin:
                raise ValueError(f"Fila {i} en 'Fases': campo 'fecha_fin' obligatorio.")

            fases.append({
                'nombre': str(nombre).strip(),
                'orden': orden,
                'descripcion': str(data.get('descripcion', '') or '').strip(),
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'presupuesto_mano_obra': data.get('presupuesto_mano_obra', 0),
                'presupuesto_materiales': data.get('presupuesto_materiales', 0),
                'presupuesto_subcontratos': data.get('presupuesto_subcontratos', 0),
                'presupuesto_equipos': data.get('presupuesto_equipos', 0),
                'presupuesto_otros': data.get('presupuesto_otros', 0),
            })

        return fases

    @staticmethod
    def _read_tareas_data(wb) -> list:
        """Lee tareas desde la hoja 'Tareas'."""
        sheet_name = 'Tareas'
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        tareas  = []

        for i, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}

            nombre = data.get('nombre')
            if not nombre:
                continue

            fase_orden = data.get('fase_orden')
            try:
                fase_orden = int(fase_orden) if fase_orden is not None else 1
            except (ValueError, TypeError):
                fase_orden = 1

            prioridad = data.get('prioridad', 2)
            try:
                prioridad = int(prioridad) if prioridad is not None else 2
                if prioridad not in (1, 2, 3, 4):
                    prioridad = 2
            except (ValueError, TypeError):
                prioridad = 2

            tareas.append({
                'nombre': str(nombre).strip(),
                'fase_orden': fase_orden,
                'descripcion': str(data.get('descripcion', '') or '').strip(),
                'fecha_inicio': ProyectoImportService._parse_date(data.get('fecha_inicio')),
                'fecha_fin': ProyectoImportService._parse_date(data.get('fecha_fin')),
                'prioridad': prioridad,
                'horas_estimadas': data.get('horas_estimadas', 0),
            })

        return tareas

    @staticmethod
    def _read_dependencias_data(wb) -> list:
        """Lee dependencias desde la hoja 'Dependencias' (opcional)."""
        sheet_name = 'Dependencias'
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return []

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        deps    = []

        for row in rows[1:]:
            if not any(row):
                continue
            data = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}

            pred_nombre = data.get('pred_nombre')
            succ_nombre = data.get('succ_nombre')
            if not pred_nombre or not succ_nombre:
                continue

            deps.append({
                'pred_nombre': str(pred_nombre).strip(),
                'succ_nombre': str(succ_nombre).strip(),
                'tipo': str(data.get('tipo', 'FS') or 'FS').strip().upper(),
                'lag': data.get('lag', 0),
            })

        return deps

    @staticmethod
    def _parse_date(value):
        """
        Intenta parsear un valor como fecha.
        Soporta: datetime objects, date objects, y strings %Y-%m-%d / %d/%m/%Y / %m/%d/%Y.
        Retorna un objeto date o None si no se puede parsear.
        """
        if value is None:
            return None

        import datetime as dt

        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value

        value_str = str(value).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
            try:
                return dt.datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue

        return None
