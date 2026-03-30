"""
SaiSuite — Proyectos: Analytics Views
Las views SOLO orquestan: reciben request → llaman service → retornan response.

Analytics endpoints:
  GET  /api/v1/projects/{pk}/analytics/kpis/
  GET  /api/v1/projects/{pk}/analytics/task-distribution/
  GET  /api/v1/projects/{pk}/analytics/velocity/
  GET  /api/v1/projects/{pk}/analytics/burn-rate/
  GET  /api/v1/projects/{pk}/analytics/burn-down/
  GET  /api/v1/projects/{pk}/analytics/resource-utilization/
  GET  /api/v1/projects/{pk}/analytics/timeline/
  POST /api/v1/projects/analytics/compare/
  POST /api/v1/projects/analytics/export-excel/
"""
import logging
from datetime import date

import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.proyectos.analytics_serializers import (
    BurnDownResponseSerializer,
    BurnRateResponseSerializer,
    CompareProjectsRequestSerializer,
    ExportExcelRequestSerializer,
    ProjectComparisonSerializer,
    ProjectKPIsSerializer,
    ProjectTimelineSerializer,
    ResourceUtilizationSerializer,
    TaskDistributionSerializer,
    VelocityResponseSerializer,
)
from apps.proyectos.analytics_services import (
    compare_projects,
    get_burn_down_data,
    get_burn_rate_data,
    get_project_kpis,
    get_project_timeline,
    get_resource_utilization,
    get_task_distribution,
    get_velocity_data,
)
from apps.proyectos.models import Project
from apps.proyectos.permissions import CanAccessProyectos

logger = logging.getLogger(__name__)


def _parse_date(value: str, field_name: str) -> date:
    """
    Parsea una cadena ISO 8601 (YYYY-MM-DD) a date.
    Levanta ValueError con mensaje claro si el formato es inválido.
    """
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        raise ValueError(
            f"El parámetro '{field_name}' debe tener formato YYYY-MM-DD. Recibido: '{value}'"
        )


def _get_project_for_company(project_pk: str, company) -> Project:
    """
    Recupera un proyecto verificando que pertenezca a la empresa del usuario.
    Retorna 404 si no existe o no pertenece a la empresa.
    """
    return get_object_or_404(Project, id=project_pk, company=company, activo=True)


# ---------------------------------------------------------------------------
# AN-01: ProjectKPIsView
# ---------------------------------------------------------------------------

class ProjectKPIsView(APIView):
    """
    KPIs principales de un proyecto.
    GET /api/v1/projects/{project_pk}/analytics/kpis/
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        data = get_project_kpis(
            project_id=str(project.id),
            company_id=str(project.company_id),
        )
        serializer = ProjectKPIsSerializer(data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-02: ProjectTaskDistributionView
# ---------------------------------------------------------------------------

class ProjectTaskDistributionView(APIView):
    """
    Distribución de tareas por estado de un proyecto.
    GET /api/v1/projects/{project_pk}/analytics/task-distribution/
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        data = get_task_distribution(
            project_id=str(project.id),
            company_id=str(project.company_id),
        )
        serializer = TaskDistributionSerializer(data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-03: ProjectVelocityView
# ---------------------------------------------------------------------------

class ProjectVelocityView(APIView):
    """
    Velocidad semanal de un proyecto (tareas completadas por semana).
    GET /api/v1/projects/{project_pk}/analytics/velocity/?periods=8
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        try:
            periods = int(request.query_params.get('periods', 8))
            if periods < 1 or periods > 52:
                return Response(
                    {'detail': 'El parámetro periods debe estar entre 1 y 52.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except (ValueError, TypeError):
            return Response(
                {'detail': 'El parámetro periods debe ser un número entero.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data_points = get_velocity_data(
            project_id=str(project.id),
            company_id=str(project.company_id),
            periods=periods,
        )
        response_data = {
            'periods': periods,
            'data': data_points,
        }
        serializer = VelocityResponseSerializer(response_data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-04: ProjectBurnRateView
# ---------------------------------------------------------------------------

class ProjectBurnRateView(APIView):
    """
    Burn rate semanal de un proyecto (horas registradas por semana).
    GET /api/v1/projects/{project_pk}/analytics/burn-rate/?periods=8
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        try:
            periods = int(request.query_params.get('periods', 8))
            if periods < 1 or periods > 52:
                return Response(
                    {'detail': 'El parámetro periods debe estar entre 1 y 52.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except (ValueError, TypeError):
            return Response(
                {'detail': 'El parámetro periods debe ser un número entero.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data_points = get_burn_rate_data(
            project_id=str(project.id),
            company_id=str(project.company_id),
            periods=periods,
        )
        response_data = {
            'periods': periods,
            'data': data_points,
        }
        serializer = BurnRateResponseSerializer(response_data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-05: ProjectBurnDownView
# ---------------------------------------------------------------------------

class ProjectBurnDownView(APIView):
    """
    Gráfico Burn Down de un proyecto.
    GET /api/v1/projects/{project_pk}/analytics/burn-down/?granularity=week
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        granularity = request.query_params.get('granularity', 'week')
        if granularity not in ('week',):
            return Response(
                {'detail': "El parámetro granularity solo acepta el valor 'week'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = get_burn_down_data(
            project_id=str(project.id),
            company_id=str(project.company_id),
            granularity=granularity,
        )
        serializer = BurnDownResponseSerializer(data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-06: ProjectResourceUtilizationView
# ---------------------------------------------------------------------------

class ProjectResourceUtilizationView(APIView):
    """
    Utilización de recursos de un proyecto.
    GET /api/v1/projects/{project_pk}/analytics/resource-utilization/
    Query params opcionales: start_date, end_date (formato YYYY-MM-DD)
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        start_date = None
        end_date = None

        start_str = request.query_params.get('start_date')
        end_str   = request.query_params.get('end_date')

        try:
            if start_str:
                start_date = _parse_date(start_str, 'start_date')
            if end_str:
                end_date = _parse_date(end_str, 'end_date')
        except ValueError as exc:
            return Response(
                {'detail': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = get_resource_utilization(
            project_id=str(project.id),
            company_id=str(project.company_id),
            start_date=start_date,
            end_date=end_date,
        )
        serializer = ResourceUtilizationSerializer(data, many=True)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-07: ProjectTimelineView
# ---------------------------------------------------------------------------

class ProjectTimelineView(APIView):
    """
    Timeline completo de un proyecto con fases y tareas.
    GET /api/v1/projects/{project_pk}/analytics/timeline/
    """
    permission_classes = [CanAccessProyectos]

    def get(self, request, project_pk):
        company = getattr(request.user, 'effective_company', None)
        project = _get_project_for_company(project_pk, company)

        data = get_project_timeline(
            project_id=str(project.id),
            company_id=str(project.company_id),
        )
        serializer = ProjectTimelineSerializer(data)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-08: CompareProjectsView
# ---------------------------------------------------------------------------

class CompareProjectsView(APIView):
    """
    Comparación de múltiples proyectos por sus métricas.
    POST /api/v1/projects/analytics/compare/
    Body: {"project_ids": ["uuid1", "uuid2", ...]}
    """
    permission_classes = [CanAccessProyectos]

    def post(self, request):
        company = getattr(request.user, 'effective_company', None)

        request_serializer = CompareProjectsRequestSerializer(data=request.data)
        if not request_serializer.is_valid():
            return Response(
                request_serializer.errors,
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_ids = [
            str(pid) for pid in request_serializer.validated_data['project_ids']
        ]
        data = compare_projects(
            project_ids=project_ids,
            company_id=str(company.id),
        )
        serializer = ProjectComparisonSerializer(data, many=True)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# AN-09: ExportExcelView
# ---------------------------------------------------------------------------

class ExportExcelView(APIView):
    """
    Exporta un reporte de analytics a Excel usando openpyxl.
    POST /api/v1/projects/analytics/export-excel/
    Body: {"project_ids": [...], "metrics": [...], "date_range": {...}}
    """
    permission_classes = [CanAccessProyectos]

    def post(self, request):
        company = getattr(request.user, 'effective_company', None)

        request_serializer = ExportExcelRequestSerializer(data=request.data)
        if not request_serializer.is_valid():
            return Response(
                request_serializer.errors,
                status=status.HTTP_400_BAD_REQUEST,
            )

        project_ids = [
            str(pid) for pid in request_serializer.validated_data['project_ids']
        ]

        logger.info(
            'analytics_excel_export',
            extra={
                'company_id': str(company.id),
                'project_ids_count': len(project_ids),
                'user_id': str(request.user.id),
            },
        )

        return _generate_excel_response(project_ids, str(company.id))


def _generate_excel_response(project_ids: list, company_id: str) -> HttpResponse:
    """
    Genera el archivo Excel con el reporte de analytics y retorna un HttpResponse.

    Incluye tres hojas:
    1. Summary — comparación general de proyectos
    2. KPIs — detalle de KPIs por proyecto
    3. Task Distribution — distribución de tareas por estado

    Args:
        project_ids: lista de UUIDs de proyectos
        company_id: UUID de la empresa (multi-tenant)

    Returns:
        HttpResponse con el archivo XLSX adjunto
    """
    wb = openpyxl.Workbook()

    # ── Hoja 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    summary_headers = [
        'Project Code',
        'Project Name',
        'Completion %',
        'On Time %',
        'Budget Variance %',
        'Velocity (tasks/week)',
        'Total Tasks',
        'Completed Tasks',
        'Overdue Tasks',
    ]
    ws_summary.append(summary_headers)

    # Aplicar negrita a cabeceras
    for cell in ws_summary[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    comparisons = compare_projects(project_ids, company_id)
    for item in comparisons:
        ws_summary.append([
            item['project_code'],
            item['project_name'],
            item['completion_rate'],
            item['on_time_rate'],
            item['budget_variance'],
            item['velocity'],
            item['total_tasks'],
            item['completed_tasks'],
            item['overdue_tasks'],
        ])

    # Ajustar ancho de columnas automáticamente
    for column in ws_summary.columns:
        max_length = max(
            (len(str(cell.value or '')) for cell in column),
            default=10,
        )
        ws_summary.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    # ── Hoja 2: KPIs ─────────────────────────────────────────────────────────
    ws_kpis = wb.create_sheet('KPIs')
    kpi_headers = [
        'Project Code',
        'Project Name',
        'Total Tasks',
        'Completed Tasks',
        'Overdue Tasks',
        'Completion Rate %',
        'On Time Rate %',
        'Budget Variance %',
        'Velocity',
        'Burn Rate (hrs/week)',
    ]
    ws_kpis.append(kpi_headers)
    for cell in ws_kpis[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Obtener proyectos válidos para la empresa
    from apps.proyectos.models import Project
    projects = Project.objects.filter(
        id__in=project_ids,
        company_id=company_id,
        activo=True,
    ).values('id', 'codigo', 'nombre')

    for project in projects:
        try:
            kpis = get_project_kpis(str(project['id']), company_id)
            ws_kpis.append([
                project['codigo'],
                project['nombre'],
                kpis['total_tasks'],
                kpis['completed_tasks'],
                kpis['overdue_tasks'],
                kpis['completion_rate'],
                kpis['on_time_rate'],
                kpis['budget_variance'],
                kpis['velocity'],
                kpis['burn_rate'],
            ])
        except Exception as exc:
            logger.warning(
                'analytics_excel_kpi_error',
                extra={
                    'project_id': str(project['id']),
                    'company_id': company_id,
                    'error': str(exc),
                },
            )

    for column in ws_kpis.columns:
        max_length = max(
            (len(str(cell.value or '')) for cell in column),
            default=10,
        )
        ws_kpis.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    # ── Hoja 3: Task Distribution ─────────────────────────────────────────────
    ws_dist = wb.create_sheet('Task Distribution')
    dist_headers = [
        'Project Code',
        'Project Name',
        'To Do',
        'In Progress',
        'In Review',
        'Completed',
        'Blocked',
        'Cancelled',
        'Total',
    ]
    ws_dist.append(dist_headers)
    for cell in ws_dist[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for project in projects:
        try:
            dist = get_task_distribution(str(project['id']), company_id)
            ws_dist.append([
                project['codigo'],
                project['nombre'],
                dist['todo'],
                dist['in_progress'],
                dist['in_review'],
                dist['completed'],
                dist['blocked'],
                dist['cancelled'],
                dist['total'],
            ])
        except Exception as exc:
            logger.warning(
                'analytics_excel_dist_error',
                extra={
                    'project_id': str(project['id']),
                    'company_id': company_id,
                    'error': str(exc),
                },
            )

    for column in ws_dist.columns:
        max_length = max(
            (len(str(cell.value or '')) for cell in column),
            default=10,
        )
        ws_dist.column_dimensions[column[0].column_letter].width = min(max_length + 4, 40)

    # ── Construir respuesta HTTP ──────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="analytics_report.xlsx"'
    wb.save(response)
    return response
