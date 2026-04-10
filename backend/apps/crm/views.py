"""
SaiSuite — CRM Views
Las views SOLO orquestan: request → service → response.
"""
import logging
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter

from apps.core.pagination import StandardPagination
from .models import CrmPipeline, CrmEtapa, CrmLead, CrmLeadScoringRule, CrmOportunidad, CrmActividad
from .serializers import (
    CrmPipelineListSerializer, CrmPipelineDetailSerializer, CrmPipelineCreateSerializer,
    CrmEtapaSerializer, CrmEtapaCreateSerializer, CrmReordenarEtapasSerializer,
    CrmLeadListSerializer, CrmLeadDetailSerializer, CrmLeadCreateSerializer,
    CrmLeadConvertirSerializer, CrmLeadScoringRuleSerializer,
    CrmOportunidadListSerializer, CrmOportunidadDetailSerializer, CrmOportunidadCreateSerializer,
    CrmMoverEtapaSerializer, CrmPerderSerializer, CrmEnviarEmailSerializer,
    CrmActividadSerializer, CrmActividadCreateSerializer, CrmCompletarActividadSerializer,
    CrmActividadAgendaSerializer,
    CrmTimelineEventSerializer, CrmNotaSerializer, CrmKanbanEtapaSerializer,
)
from .services import (
    PipelineService, EtapaService, LeadService, LeadScoringService,
    OportunidadService, ActividadService, TimelineService,
)
from .permissions import CanAccessCrm, CrmBasePermission, CrmAdminPermission, CrmImportPermission
from .filters import CrmLeadFilter, CrmOportunidadFilter, CrmActividadFilter

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# PIPELINE
# ─────────────────────────────────────────────

class PipelineViewSet(viewsets.ViewSet):
    permission_classes = [CrmAdminPermission]

    def list(self, request):
        pipelines = PipelineService.list(request.user.company)
        return Response(CrmPipelineListSerializer(pipelines, many=True).data)

    def retrieve(self, request, pk=None):
        pipeline = PipelineService.get(pk, request.user.company)
        return Response(CrmPipelineDetailSerializer(pipeline).data)

    def create(self, request):
        ser = CrmPipelineCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        pipeline = PipelineService.create(request.user.company, ser.validated_data)
        return Response(CrmPipelineDetailSerializer(pipeline).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, pk=None):
        pipeline = PipelineService.get(pk, request.user.company)
        ser = CrmPipelineCreateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        pipeline = PipelineService.update(pipeline, ser.validated_data)
        return Response(CrmPipelineDetailSerializer(pipeline).data)

    def destroy(self, request, pk=None):
        pipeline = PipelineService.get(pk, request.user.company)
        try:
            PipelineService.delete(pipeline)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'], permission_classes=[CanAccessCrm])
    def kanban(self, request, pk=None):
        """Vista Kanban: oportunidades agrupadas por etapa."""
        pipeline = PipelineService.get(pk, request.user.company)
        kanban = PipelineService.get_kanban(
            pipeline,
            asignado_a=request.query_params.get('asignado_a'),
            search=request.query_params.get('search', ''),
            page=int(request.query_params.get('page', 1)),
            page_size=int(request.query_params.get('page_size', 50)),
        )
        return Response(CrmKanbanEtapaSerializer(kanban, many=True).data)


# ─────────────────────────────────────────────
# ETAPAS
# ─────────────────────────────────────────────

class EtapaViewSet(viewsets.ViewSet):
    permission_classes = [CrmAdminPermission]

    def list(self, request, pipeline_pk=None):
        pipeline = PipelineService.get(pipeline_pk, request.user.company)
        etapas = EtapaService.list(pipeline)
        return Response(CrmEtapaSerializer(etapas, many=True).data)

    def create(self, request, pipeline_pk=None):
        pipeline = PipelineService.get(pipeline_pk, request.user.company)
        ser = CrmEtapaCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        etapa = EtapaService.create(pipeline, ser.validated_data)
        return Response(CrmEtapaSerializer(etapa).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, pk=None, pipeline_pk=None):
        etapa = EtapaService.get(pk, request.user.company)
        ser = CrmEtapaCreateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        etapa = EtapaService.update(etapa, ser.validated_data)
        return Response(CrmEtapaSerializer(etapa).data)

    def destroy(self, request, pk=None, pipeline_pk=None):
        etapa = EtapaService.get(pk, request.user.company)
        try:
            EtapaService.delete(etapa)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='reordenar')
    def reordenar(self, request, pipeline_pk=None):
        pipeline = PipelineService.get(pipeline_pk, request.user.company)
        ser = CrmReordenarEtapasSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        EtapaService.reordenar(pipeline, [str(i) for i in ser.validated_data['orden']])
        return Response({'status': 'ok'})


# ─────────────────────────────────────────────
# LEADS
# ─────────────────────────────────────────────

class LeadViewSet(viewsets.ViewSet):
    permission_classes = [CrmBasePermission]
    filter_backends    = [DjangoFilterBackend]
    filterset_class    = CrmLeadFilter

    def list(self, request):
        qs = LeadService.list(
            request.user.company,
            search=request.query_params.get('search', ''),
            fuente=request.query_params.get('fuente', ''),
            convertido=_parse_bool(request.query_params.get('convertido')),
            asignado_a=request.query_params.get('asignado_a'),
            pipeline_id=request.query_params.get('pipeline'),
        )
        paginator = StandardPagination()
        page = paginator.paginate_queryset(qs, request)
        return paginator.get_paginated_response(CrmLeadListSerializer(page, many=True).data)

    def retrieve(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        return Response(CrmLeadDetailSerializer(lead).data)

    def create(self, request):
        ser = CrmLeadCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        pipeline_id = data.pop('pipeline_id', None)
        if pipeline_id:
            from .models import CrmPipeline
            data['pipeline'] = CrmPipeline.objects.filter(id=pipeline_id, company=request.user.company).first()
        lead = LeadService.create(request.user.company, data)
        return Response(CrmLeadDetailSerializer(lead).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        self.check_object_permissions(request, lead)
        ser = CrmLeadCreateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        lead = LeadService.update(lead, ser.validated_data)
        return Response(CrmLeadDetailSerializer(lead).data)

    def destroy(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        self.check_object_permissions(request, lead)
        try:
            LeadService.delete(lead)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def convertir(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        ser = CrmLeadConvertirSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            oportunidad = LeadService.convertir(lead, ser.validated_data)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CrmOportunidadDetailSerializer(oportunidad).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def asignar(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'detail': 'user_id requerido'}, status=status.HTTP_400_BAD_REQUEST)
        lead = LeadService.asignar(lead, user_id)
        return Response(CrmLeadDetailSerializer(lead).data)

    @action(detail=True, methods=['get'], url_path='actividades')
    def list_actividades(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        solo_pendientes = _parse_bool(request.query_params.get('solo_pendientes'))
        qs = ActividadService.list_for_lead(lead, solo_pendientes=solo_pendientes or False)
        return Response(CrmActividadSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='actividades')
    def crear_actividad(self, request, pk=None):
        lead = LeadService.get(pk, request.user.company)
        ser = CrmActividadCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        asignado_id = data.pop('asignado_a_id', None)
        if asignado_id:
            from apps.users.models import User
            data['asignado_a'] = User.objects.filter(id=asignado_id, company=request.user.company).first()
        actividad = ActividadService.create_for_lead(lead, data)
        return Response(CrmActividadSerializer(actividad).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='round-robin')
    def round_robin(self, request, pk=None):
        """Asigna el lead al vendedor con menos leads activos."""
        try:
            lead = LeadService.get(pk, request.user.company)
        except Exception:
            return Response({'detail': 'Lead no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        lead = LeadService.asignar_round_robin(lead)
        return Response(CrmLeadDetailSerializer(lead).data)

    @action(detail=False, methods=['post'], url_path='asignar-masivo')
    def asignar_masivo(self, request):
        """Asigna vía round-robin todos los leads sin asignado de la empresa."""
        asignados = LeadService.asignar_masivo_round_robin(request.user.company)
        return Response({'asignados': asignados})

    @action(detail=False, methods=['post'], permission_classes=[CrmImportPermission],
            parser_classes=[MultiPartParser])
    def importar(self, request):
        """Importa leads desde CSV/Excel."""
        import openpyxl
        from io import BytesIO

        archivo = request.FILES.get('archivo')
        pipeline_id = request.data.get('pipeline_id')
        if not archivo:
            return Response({'detail': 'Se requiere un archivo'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(BytesIO(archivo.read()))
            ws = wb.active
            headers = [str(c.value).lower().strip() if c.value else '' for c in ws[1]]
            registros = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                fila = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}
                registros.append(fila)
        except Exception as e:
            return Response({'detail': f'Error al leer archivo: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        resultado = LeadService.importar_csv(request.user.company, registros, pipeline_id)
        return Response(resultado, status=status.HTTP_201_CREATED)


class LeadWebhookView(APIView):
    """Endpoint público para captura de leads desde formularios web."""
    authentication_classes = []
    permission_classes     = []

    def post(self, request, company_nit=None):
        from apps.companies.models import Company

        try:
            company = Company.objects.get(nit=company_nit)
        except Company.DoesNotExist:
            return Response({'detail': 'Empresa no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        lead = LeadService.create_from_webhook(company, request.data)
        return Response({'id': str(lead.id), 'score': lead.score}, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────
# LEAD SCORING RULES
# ─────────────────────────────────────────────

class LeadScoringRuleViewSet(viewsets.ViewSet):
    permission_classes = [CrmAdminPermission]

    def list(self, request):
        reglas = CrmLeadScoringRule.objects.filter(company=request.user.company)
        return Response(CrmLeadScoringRuleSerializer(reglas, many=True).data)

    def create(self, request):
        ser = CrmLeadScoringRuleSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        regla = CrmLeadScoringRule.objects.create(
            company=request.user.company, **ser.validated_data
        )
        return Response(CrmLeadScoringRuleSerializer(regla).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, pk=None):
        regla = CrmLeadScoringRule.objects.get(id=pk, company=request.user.company)
        ser = CrmLeadScoringRuleSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        for key, value in ser.validated_data.items():
            setattr(regla, key, value)
        regla.save()
        return Response(CrmLeadScoringRuleSerializer(regla).data)

    def destroy(self, request, pk=None):
        regla = CrmLeadScoringRule.objects.get(id=pk, company=request.user.company)
        regla.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─────────────────────────────────────────────
# OPORTUNIDADES
# ─────────────────────────────────────────────

class OportunidadViewSet(viewsets.ViewSet):
    permission_classes = [CrmBasePermission]

    def list(self, request):
        qs = OportunidadService.list(
            request.user.company,
            search=request.query_params.get('search', ''),
            pipeline_id=request.query_params.get('pipeline'),
            etapa_id=request.query_params.get('etapa'),
            asignado_a=request.query_params.get('asignado_a'),
            ganada=_parse_bool(request.query_params.get('ganada')),
            perdida=_parse_bool(request.query_params.get('perdida')),
        )
        paginator = StandardPagination()
        page = paginator.paginate_queryset(qs, request)
        return paginator.get_paginated_response(CrmOportunidadListSerializer(page, many=True).data)

    def retrieve(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        return Response(CrmOportunidadDetailSerializer(op).data)

    def create(self, request):
        ser = CrmOportunidadCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        op = OportunidadService.create(request.user.company, ser.validated_data)
        return Response(CrmOportunidadDetailSerializer(op).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        self.check_object_permissions(request, op)
        ser = CrmOportunidadCreateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        op = OportunidadService.update(op, ser.validated_data)
        return Response(CrmOportunidadDetailSerializer(op).data)

    def destroy(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        self.check_object_permissions(request, op)
        OportunidadService.delete(op)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='mover-etapa')
    def mover_etapa(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        ser = CrmMoverEtapaSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            op = OportunidadService.mover_etapa(op, str(ser.validated_data['etapa_id']), usuario=request.user)
        except (CrmEtapa.DoesNotExist, ValueError) as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CrmOportunidadDetailSerializer(op).data)

    @action(detail=True, methods=['post'])
    def ganar(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        try:
            op = OportunidadService.ganar(op, usuario=request.user)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CrmOportunidadDetailSerializer(op).data)

    @action(detail=True, methods=['post'])
    def perder(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        ser = CrmPerderSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            op = OportunidadService.perder(op, ser.validated_data['motivo'], usuario=request.user)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CrmOportunidadDetailSerializer(op).data)

    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        eventos = TimelineService.list(op)
        return Response(CrmTimelineEventSerializer(eventos, many=True).data)

    @action(detail=True, methods=['post'], url_path='timeline')
    def agregar_nota(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        ser = CrmNotaSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        evento = TimelineService.agregar_nota(op, ser.validated_data['nota'], usuario=request.user)
        return Response(CrmTimelineEventSerializer(evento).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='enviar-email')
    def enviar_email(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        ser = CrmEnviarEmailSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            OportunidadService.enviar_email(op, ser.validated_data['asunto'], ser.validated_data['cuerpo'], usuario=request.user)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'status': 'email enviado'})

    # ── Actividades nested ────────────────────

    @action(detail=True, methods=['get'], url_path='actividades')
    def list_actividades(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        qs = ActividadService.list(op, solo_pendientes=_parse_bool(request.query_params.get('solo_pendientes')))
        return Response(CrmActividadSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='actividades')
    def crear_actividad(self, request, pk=None):
        op = OportunidadService.get(pk, request.user.company)
        ser = CrmActividadCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        asignado_id = data.pop('asignado_a_id', None)
        if asignado_id:
            from apps.users.models import User
            data['asignado_a'] = User.objects.filter(id=asignado_id, company=request.user.company).first()
        actividad = ActividadService.create(op, data)
        return Response(CrmActividadSerializer(actividad).data, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────
# ACTIVIDADES (endpoints individuales)
# ─────────────────────────────────────────────

class ActividadViewSet(viewsets.ViewSet):
    permission_classes = [CrmBasePermission]

    def retrieve(self, request, pk=None):
        actividad = ActividadService.get(pk, request.user.company)
        return Response(CrmActividadSerializer(actividad).data)

    def partial_update(self, request, pk=None):
        actividad = ActividadService.get(pk, request.user.company)
        ser = CrmActividadCreateSerializer(data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        actividad = ActividadService.update(actividad, ser.validated_data)
        return Response(CrmActividadSerializer(actividad).data)

    def destroy(self, request, pk=None):
        actividad = ActividadService.get(pk, request.user.company)
        ActividadService.delete(actividad)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def completar(self, request, pk=None):
        actividad = ActividadService.get(pk, request.user.company)
        ser = CrmCompletarActividadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        actividad = ActividadService.completar(actividad, ser.validated_data.get('resultado', ''), usuario=request.user)
        return Response(CrmActividadSerializer(actividad).data)


# ─────────────────────────────────────────────
# AGENDA (actividades globales de la empresa)
# ─────────────────────────────────────────────

class AgendaView(APIView):
    """Actividades de toda la empresa (oportunidades + leads), filtradas por rango de fechas."""
    permission_classes = [CanAccessCrm]

    def get(self, request):
        from django.utils.dateparse import parse_date
        from django.db.models import Q
        company = request.user.company

        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')
        solo_pendientes = _parse_bool(request.query_params.get('solo_pendientes'))
        asignado_a = request.query_params.get('asignado_a')

        qs = CrmActividad.all_objects.filter(
            Q(oportunidad__company=company) | Q(lead__company=company)
        ).select_related('asignado_a', 'oportunidad', 'lead')

        if fecha_desde:
            d = parse_date(fecha_desde)
            if d:
                qs = qs.filter(fecha_programada__date__gte=d)
        if fecha_hasta:
            d = parse_date(fecha_hasta)
            if d:
                qs = qs.filter(fecha_programada__date__lte=d)
        if solo_pendientes:
            qs = qs.filter(completada=False)
        if asignado_a:
            qs = qs.filter(asignado_a__id=asignado_a)

        qs = qs.order_by('fecha_programada')
        return Response(CrmActividadAgendaSerializer(qs, many=True).data)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _parse_bool(value):
    if value is None:
        return None
    return str(value).lower() in ('true', '1', 'yes')
